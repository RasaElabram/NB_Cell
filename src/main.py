import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import sqlite3
import re  
import openpyxl
import os.path
from openpyxl import Workbook, load_workbook

_location = os.path.dirname(__file__)

class Toplevel1:
    def __init__(self, top=None):
        self.source_dump_path = None
        self.selected_sheet = None
        self.template_path = os.path.join(_location, "..", "Templates", "template.xlsx")
        self.conn = None  # Initialize the connection attribute

        top.geometry("600x450+468+170")
        top.minsize(120, 1)
        top.maxsize(1540, 941)
        top.resizable(1,  1)
        top.title("NPO Report Assistant")
        top.configure(background="#919191")
        top.configure(highlightbackground="#919191")
        top.configure(highlightcolor="white")

        self.top = top

        self.menubar = tk.Menu(top, font="TkMenuFont", bg='#919191', fg='white')
        top.configure(menu=self.menubar)

        self.Button1 = tk.Button(self.top, command=self.connect_to_database)
        self.Button1.place(relx=0.05, rely=0.133, height=26, width=67)
        self.Button1.configure(activebackground="#d9d9d9")
        self.Button1.configure(activeforeground="black")
        self.Button1.configure(background="#919191")
        self.Button1.configure(disabledforeground="#adadad")
        self.Button1.configure(foreground="white")
        self.Button1.configure(highlightbackground="#919191")
        self.Button1.configure(highlightcolor="white")
        self.Button1.configure(text='''Update''')

        self.TSeparator1 = ttk.Separator(self.top)
        self.TSeparator1.place(relx=0.2, rely=0.0,  relheight=0.311)
        self.TSeparator1.configure(orient="vertical")

        self.TSeparator2 = ttk.Separator(self.top)
        self.TSeparator2.place(relx=0.0, rely=0.311,  relwidth=0.2)

        self.Label1 = tk.Label(self.top)
        self.Label1.place(relx=0.05, rely=0.044, height=31, width=54)
        self.Label1.configure(anchor='w')
        self.Label1.configure(background="#919191")
        self.Label1.configure(foreground="white")
        self.Label1.configure(text='''Database''')

        self.Label2 = tk.Label(self.top)
        self.Label2.place(relx=0.1, rely=0.422, height=31, width=94)
        self.Label2.configure(anchor='w')
        self.Label2.configure(background="#919191")
        self.Label2.configure(foreground="white")
        self.Label2.configure(text='''4G Parameters''')

        self.Button2 = tk.Button(self.top, command=self.upload_excel_file)
        self.Button2.place(relx=0.117, rely=0.511, height=27, width=120)
        self.Button2.configure(activebackground="#d9d9d9")
        self.Button2.configure(activeforeground="black")
        self.Button2.configure(background="#919191")
        self.Button2.configure(disabledforeground="#adadad")
        self.Button2.configure(foreground="white")
        self.Button2.configure(highlightbackground="#919191")
        self.Button2.configure(highlightcolor="white")
        self.Button2.configure(text='''Upload Source Dump''')

        self.Button3 = tk.Button(self.top, command=self.execute_source_cell)
        self.Button3.place(relx=0.117, rely=0.611, height=27, width=120)
        self.Button3.configure(activebackground="#d9d9d9")
        self.Button3.configure(activeforeground="black")
        self.Button3.configure(background="#919191")
        self.Button3.configure(disabledforeground="#adadad")
        self.Button3.configure(foreground="white")
        self.Button3.configure(highlightbackground="#919191")
        self.Button3.configure(highlightcolor="white")
        self.Button3.configure(text='''NB_EUtranCellFDDLTE''')

    def connect_to_database(self):
        try:
            # Get the directory of the current script
            script_dir = os.path.dirname(__file__)

            # Construct the path to the database file in the "Database" folder
            db_path = os.path.join(script_dir, '..', 'Database', 'cell_info.db')

            # Establish connection to the database
            self.conn = sqlite3.connect(db_path)

            messagebox.showinfo("Database Connection", "Connected to the database successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while connecting to the database: {str(e)}")

    def upload_excel_file(self):
        print("Clicked")  # Check if button click event is captured

        # Prompt the user to select an Excel file
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if not file_path:
            return  # Return if no file selected or dialog closed
        
        # Prompt the user to select a sheet from the selected Excel file
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names  # Corrected

        # Create a new window for the sheet selection
        sheet_window = tk.Toplevel(self.top)
        sheet_window.title("Select Sheet")
        
        # Calculate the center position for the sheet window
        sheet_window.geometry("300x100")

        # Create a label for the sheet selection
        label = tk.Label(sheet_window, text="Select Sheet:")
        label.pack(pady=5)

        # Create a combo box for selecting the sheet
        selected_sheet = tk.StringVar()
        combo = ttk.Combobox(sheet_window, textvariable=selected_sheet, values=sheet_names)
        combo.pack(pady=5)

        # Function to handle the selection
        def select_sheet():
            selected_sheet_name = combo.get()
            if selected_sheet_name:
                self.source_dump_path = file_path
                self.selected_sheet = selected_sheet_name  # Update the selected_sheet attribute
                sheet_window.destroy()

        # Create a button to confirm selection
        select_button = tk.Button(sheet_window, text="Select", command=select_sheet)
        select_button.pack(pady=5)

    def execute_source_cell(self):
        if not self.conn:
            messagebox.showwarning("Warning", "Please connect to the database first.")
            return
            
        if not (self.source_dump_path and self.selected_sheet):
            messagebox.showwarning("Warning", "Please upload source dump and select a sheet before executing.")
            return

        try:
            # Read data from the source dump Excel file
            source_data = pd.read_excel(self.source_dump_path, sheet_name=self.selected_sheet)
            
            # Call compare_data_and_output method with the database connection and source_data
            self.compare_data_and_output(self.conn, source_data)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while comparing data: {str(e)}")

    def compare_data_and_output(self, conn, source_data):
        try:
            # Check if the output template file exists
            output_template_path = os.path.join(_location, "..", "Templates", "output_template.xlsx")

            if os.path.exists(output_template_path):
                # Load existing workbook if the output template exists
                wb = load_workbook(output_template_path)
            else:
                # Create a new workbook if the output template doesn't exist
                wb = Workbook()
                # Add a new sheet to the workbook
                ws = wb.active
                ws.title = "Output"
                # Add headers if creating a new template
                headers = ["Header1", "Header2", "Header3"]  # Add your headers here
                ws.append(headers)

            ws = wb.active

            # Placeholder for the comparison logic with the database
            print("Results for matching rows:")

            # Flag to indicate if at least one match is found
            match_found = False

            # Start from row 6 in the source data and row 4 in the output template
            row_index_source = 2
            row_index_template = 4

            # Copy data from column K of the source data to column H of the output template
            for index, row in source_data.iterrows():
                if row_index_source >= 6:  # Start copying from row 6
                    source_cell_data = row.iloc[10]  # Get the value from column K (0-based index is 10)
                    # Write the source cell data to column H of the template
                    ws.cell(row=row_index_template, column=8, value=source_cell_data)

                    row_index_template += 1  # Increment template row index only when copying data

                row_index_source += 1  # Always increment source row index

            # Loop through the copied data in column H of the template and perform database matching
            for row_index_template in range(3, row_index_template):  # Iterate over copied data rows
                source_cell_data = ws.cell(row=row_index_template, column=8).value

                if pd.notnull(source_cell_data):  # Check if the value is not null
                    # Rest of the code for database matching remains the same...
                    source_cell_data = str(source_cell_data)  # Convert to string

                    # Use regular expression to extract numbers from the source cell data
                    numbers = re.findall(r'ENBCUCPFunction=(\d+),.*?CUEUtranCellFDDLTE=(\d+)', source_cell_data)

                    if numbers:
                        match_found = True  # Set flag to indicate match found
                        sourceEnbId, CellId = numbers[0]  # Extract the numbers from the tuple

                        # Convert to integers
                        sourceEnbId = int(sourceEnbId)
                        CellId = int(CellId)

                        # Concatenate sourceEnbId and CellId to form the concatenated value
                        concatenated_value = str(sourceEnbId) + str(CellId)

                        # Execute SQL query to find the corresponding id from the concatenated values of enbid and celllocalid
                        cursor_concatenated = conn.cursor()
                        cursor_concatenated.execute("""
                            SELECT id FROM (
                                SELECT enbid.id, enbid.eNBId || celllocalid.CellLocalId as concatenated_value 
                                FROM enbid
                                JOIN celllocalid ON enbid.id = celllocalid.id
                            ) AS concatenated_table
                            WHERE concatenated_value = ?
                        """, (concatenated_value,))
                        result_concatenated = cursor_concatenated.fetchone()

                        # If result is found, print the corresponding id
                        if result_concatenated:
                            concatenated_id = result_concatenated[0]

                            # Execute SQL query to retrieve CellName using the concatenated_id
                            cursor_cellname = conn.cursor()
                            cursor_cellname.execute("SELECT CellName FROM CellName WHERE id = ?", (concatenated_id,))
                            result_cellname = cursor_cellname.fetchone()

                            # If CellName is found, write it to the template file in column I
                            if result_cellname:
                                cell_name = result_cellname[0]
                                ws.cell(row=row_index_template, column=9, value=cell_name)  # Write to column I (index 9)
                                print(f"CellName '{cell_name}' written to template file at row {row_index_template}.")
                            else:
                                print("No CellName found for the concatenated ID.")
                        else:
                            print(f"No match found for sourceEnbId={sourceEnbId} and CellId={CellId}.")
                    else:
                        print("Source cell data does not contain the expected pattern.")

            # If no match is found, print a message
            if not match_found:
                print("No matching rows found.")

            # Save the modified template file
            wb.save(output_template_path)

            # Close the database connection
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while comparing data: {str(e)}")




def main():
    root = tk.Tk()
    Toplevel1(root)
    root.mainloop()

if __name__ == "__main__":
    main()
